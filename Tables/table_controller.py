import pandas as pd
import openpyxl
from tkinter import filedialog
from tkinter import messagebox


def open_file():
    """
    Функция открытия диалогового окна для получения пути к файлу \n
    Пример: button = Button(text="Open", command=open_file)

    :return: путь к файлу
    """
    file_path = filedialog.askopenfilename()
    return file_path


def open_folder():
    """
    Выбор папки для сохранения полученной таблицы \n
    Пример: button = Button(text="Open", command=open_folder)

    :return: путь к папке
    """
    folder_path = filedialog.askdirectory()
    return folder_path


def open_table_xlsx(file_path):
    """
    Загрузка таблицы

    :param file_path: путь к файлу .xlsx
    :return: таблица Data Frame
    """
    df = pd.read_excel(file_path)
    print(df.shape)
    df.rename(columns={'Track.Name': 'Track_name', 'Artist.Name': 'Artist_name', 'Date.of.release': 'Date_of_release',
                       'Beats.Per.Minute': 'Beats_per_minute', 'Loudness..dB..': 'Loudness',
                       'Monthly.auditions': 'Monthly_auditions', 'Auditions.on.the.track': 'Auditions_on_the_track'},
              inplace=True)
    print(df.dtypes)
    return df


def save_dataframe_excel(dataframe, folder_path, name_of_report):
    """
    Сохранение DataFrame в .xlsx таблицу по выбранному пути

    :param dataframe: созданный Data Frame
    :param folder_path: путь к папке сохранения
    :param name_of_report: название сохраняемого файла
    """
    path_to_save = folder_path + name_of_report + '.xlsx'
    writer = pd.ExcelWriter(path_to_save)
    dataframe.to_excel(writer)
    writer.save()


def add_data_to_table(file_path, data):
    """
    Добавление новой строчки в файл xlsx \n
    Вид data - [№, Track.Name, Artist.Name, Collaboration, Genre, Country, Date.of.release, Album, Beats.Per.Minute,
    Loudness..dB.., Length, Popularity, Monthly.auditions, Auditions.on.the.track, Label]

    :param file_path: путь к файлу
    :param data: добавляемая строчка
    """
    if data.count('') > 0:
        messagebox.showerror('Ошибка', 'Не все поля заполнены!')
        return
    data_ = []
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    data_.append(ws.max_row + 1)
    for i in range(len(data)):
        data_.append(data[i])
    ws.append(data_)
    wb.save(file_path)
    messagebox.showinfo('Успех', 'Изменения успешно внесены!\n Для отображения изменений вернитесь в главное меню'
                                 ' и нажмите кнопку Начать')


def delete_data_from_table(file_path, index):
    """
    Удаляет выбранную строчку из таблицы

    :param file_path: путь к файлу .xlsx
    :param index: удаляемая строка (нумерация как в excel таблице)
    """
    if index == '':
        messagebox.showerror('Ошибка', 'Индекс не выбран!')
        return
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    rows = ws.max_row
    ws.delete_rows(idx=int(index), amount=1)
    wb.save(file_path)
    messagebox.showinfo('Успех', 'Изменения успешно внесены!\n Для отображения изменений вернитесь в главное меню'
                                 ' и нажмите кнопку Начать')


def edit_data_in_table(file_path, new_data, cell):
    """
    Изменение текущей ячейки

    :param file_path: путь к файлу .xlsx
    :param new_data: новое содержание ячейки
    :param cell: изменяемая ячейка (Буква + цифра. Пример 'A4')
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws[cell] = new_data
    ws.save(file_path)